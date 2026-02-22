"""Unit tests for tenancy_parser module."""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from ocr_extractor.table_detector import CellRegion, TableGrid
from ocr_extractor.tenancy_parser import (
    HTML_SCHEMA_COLUMNS,
    ROW_TYPE_CHARGE_SCHEDULE,
    ROW_TYPE_LEASE_SUMMARY,
    ROW_TYPE_OCCUPANCY_SUMMARY,
    ROW_TYPE_RENT_STEP,
    TenancyRow,
    export_tenancy_to_excel,
    export_tenancy_to_html,
    normalize_date,
    normalize_number,
    parse_grid_to_rows,
    _detect_section_type,
    _extract_property_as_of_date,
)


class TestNormalizeNumber:
    """Test numeric value normalization."""

    def test_simple_integer(self):
        assert normalize_number("123") == 123.0

    def test_simple_float(self):
        assert normalize_number("123.45") == 123.45

    def test_with_commas(self):
        assert normalize_number("1,234.56") == 1234.56

    def test_with_dollar_sign(self):
        assert normalize_number("$1,234.56") == 1234.56

    def test_negative_parentheses(self):
        assert normalize_number("(100)") == -100.0

    def test_negative_parentheses_with_commas(self):
        assert normalize_number("($1,234.56)") == -1234.56

    def test_ocr_o_to_zero(self):
        # Common OCR error: O instead of 0
        assert normalize_number("1O0") == 100.0
        assert normalize_number("1o0") == 100.0

    def test_invalid_returns_none(self):
        assert normalize_number("abc") is None
        assert normalize_number("") is None
        assert normalize_number("   ") is None

    def test_none_returns_none(self):
        assert normalize_number(None) is None


class TestNormalizeDate:
    """Test date normalization to ISO format."""

    def test_mdy_slash(self):
        assert normalize_date("01/15/2024") == "2024-01-15"
        assert normalize_date("1/5/2024") == "2024-01-05"

    def test_mdy_slash_short_year(self):
        assert normalize_date("01/15/24") == "2024-01-15"

    def test_iso_format(self):
        assert normalize_date("2024-01-15") == "2024-01-15"

    def test_month_name(self):
        assert normalize_date("January 15, 2024") == "2024-01-15"
        assert normalize_date("Jan 15, 2024") == "2024-01-15"

    def test_invalid_returns_none(self):
        assert normalize_date("not-a-date") is None
        assert normalize_date("") is None
        assert normalize_date("   ") is None

    def test_none_returns_none(self):
        assert normalize_date(None) is None


class TestTenancyRow:
    """Test TenancyRow dataclass."""

    def test_to_dict_all_fields(self):
        row = TenancyRow(
            property="Test Property",
            as_of_date="2024-01-15",
            tenant_name="Test Tenant",
            legal_name="Test Tenant LLC",
            suite="101",
            lease_type="Gross",
            lease_from="2024-01-01",
            lease_to="2024-12-31",
            term_months=12.0,
            area_sqft=1500.0,
            monthly_amount=1000.0,
            annual_amount=12000.0,
            security_deposit=2000.0,
            loc_amount=5000.0,
            notes="Test notes",
            row_type="lease_summary",
        )
        data = row.to_dict()
        assert data["property"] == "Test Property"
        assert data["as_of_date"] == "2024-01-15"
        assert data["tenant_name"] == "Test Tenant"
        assert data["legal_name"] == "Test Tenant LLC"
        assert data["suite"] == "101"
        assert data["lease_type"] == "Gross"
        assert data["lease_from"] == "2024-01-01"
        assert data["lease_to"] == "2024-12-31"
        assert data["term_months"] == 12.0
        assert data["area_sqft"] == 1500.0
        assert data["monthly_amount"] == 1000.0
        assert data["annual_amount"] == 12000.0
        assert data["security_deposit"] == 2000.0
        assert data["loc_amount"] == 5000.0
        assert data["notes"] == "Test notes"
        assert data["row_type"] == "lease_summary"

    def test_to_dict_with_warnings(self):
        row = TenancyRow(
            property="Test",
            warnings=["Warning 1", "Warning 2"],
        )
        data = row.to_dict()
        assert data["warnings"] == "Warning 1; Warning 2"

    def test_to_dict_no_warnings(self):
        row = TenancyRow(property="Test")
        data = row.to_dict()
        assert data["warnings"] is None


class TestParseGridToRows:
    """Test grid-to-rows parsing."""

    def test_empty_grid(self):
        grid = TableGrid()
        rows = parse_grid_to_rows(grid)
        assert rows == []

    def test_single_row_header_only(self):
        grid = TableGrid(
            cells=[
                CellRegion(row=0, col=0, text="Property"),
                CellRegion(row=0, col=1, text="Tenant"),
                CellRegion(row=0, col=2, text="Suite"),
            ],
            header_rows=1,
        )
        rows = parse_grid_to_rows(grid)
        assert rows == []  # No data rows

    def test_header_and_data_row(self):
        grid = TableGrid(
            cells=[
                # Header
                CellRegion(row=0, col=0, text="Property"),
                CellRegion(row=0, col=1, text="Tenant"),
                CellRegion(row=0, col=2, text="Suite"),
                # Data
                CellRegion(row=1, col=0, text="Building A"),
                CellRegion(row=1, col=1, text="ABC Corp"),
                CellRegion(row=1, col=2, text="101"),
            ],
            header_rows=1,
        )
        rows = parse_grid_to_rows(grid)
        assert len(rows) == 1
        assert rows[0].property == "Building A"
        assert rows[0].tenant_name == "ABC Corp"
        assert rows[0].suite == "101"

    def test_numeric_field_normalization(self):
        grid = TableGrid(
            cells=[
                # Header
                CellRegion(row=0, col=0, text="Monthly"),
                CellRegion(row=0, col=1, text="Annual"),
                # Data with commas
                CellRegion(row=1, col=0, text="$1,000.50"),
                CellRegion(row=1, col=1, text="$12,006.00"),
            ],
            header_rows=1,
        )
        rows = parse_grid_to_rows(grid)
        assert len(rows) == 1
        assert rows[0].monthly_amount == 1000.50
        assert rows[0].annual_amount == 12006.00

    def test_date_field_normalization(self):
        grid = TableGrid(
            cells=[
                # Header
                CellRegion(row=0, col=0, text="Lease From"),
                CellRegion(row=0, col=1, text="Lease To"),
                # Data
                CellRegion(row=1, col=0, text="01/01/2024"),
                CellRegion(row=1, col=1, text="12/31/2024"),
            ],
            header_rows=1,
        )
        rows = parse_grid_to_rows(grid)
        assert len(rows) == 1
        assert rows[0].lease_from == "2024-01-01"
        assert rows[0].lease_to == "2024-12-31"

    def test_invalid_numeric_adds_warning(self):
        grid = TableGrid(
            cells=[
                # Header
                CellRegion(row=0, col=0, text="Property"),
                CellRegion(row=0, col=1, text="Monthly"),
                # Data with property (so it's not filtered) and invalid number
                CellRegion(row=1, col=0, text="Building A"),
                CellRegion(row=1, col=1, text="invalid"),
            ],
            header_rows=1,
        )
        rows = parse_grid_to_rows(grid)
        assert len(rows) == 1
        assert rows[0].property == "Building A"
        assert rows[0].monthly_amount is None
        assert len(rows[0].warnings) > 0
        assert "monthly_amount" in rows[0].warnings[0].lower()

    def test_invalid_date_adds_warning(self):
        grid = TableGrid(
            cells=[
                # Header
                CellRegion(row=0, col=0, text="Lease From"),
                # Data with invalid date
                CellRegion(row=1, col=0, text="invalid-date"),
            ],
            header_rows=1,
        )
        rows = parse_grid_to_rows(grid)
        assert len(rows) == 1
        assert rows[0].lease_from == "invalid-date"  # Keeps original
        assert len(rows[0].warnings) > 0
        assert "date" in rows[0].warnings[0].lower()


class TestExportTenancyToExcel:
    """Test Excel export with multi-column guarantee."""

    def test_export_creates_file(self, tmp_path):
        rows = [
            TenancyRow(
                property="Test Property",
                tenant_name="Test Tenant",
                suite="101",
            ),
        ]
        output_path = tmp_path / "test.xlsx"
        result = export_tenancy_to_excel(rows, output_path)
        assert Path(result).exists()

    def test_export_has_multiple_columns(self, tmp_path):
        rows = [
            TenancyRow(
                property="Test Property",
                tenant_name="Test Tenant",
                suite="101",
                monthly_amount=1000.0,
                annual_amount=12000.0,
            ),
        ]
        output_path = tmp_path / "test.xlsx"
        export_tenancy_to_excel(rows, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Should have at least 10 columns (schema defines 17)
        assert ws.max_column >= 10

    def test_export_header_row(self, tmp_path):
        rows = [
            TenancyRow(property="Test", tenant_name="Tenant"),
        ]
        output_path = tmp_path / "test.xlsx"
        export_tenancy_to_excel(rows, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Check header row contains expected column names
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        assert "property" in headers
        assert "tenant_name" in headers
        assert "suite" in headers
        assert "monthly_amount" in headers

    def test_export_data_in_correct_columns(self, tmp_path):
        rows = [
            TenancyRow(
                property="Building A",
                tenant_name="ABC Corp",
                suite="101",
                monthly_amount=1000.0,
            ),
        ]
        output_path = tmp_path / "test.xlsx"
        export_tenancy_to_excel(rows, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Get header row to find column indices
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

        # Find column indices
        property_col = headers.index("property") + 1
        tenant_col = headers.index("tenant_name") + 1
        suite_col = headers.index("suite") + 1
        monthly_col = headers.index("monthly_amount") + 1

        # Check data row (row 2)
        assert ws.cell(2, property_col).value == "Building A"
        assert ws.cell(2, tenant_col).value == "ABC Corp"
        assert ws.cell(2, suite_col).value == "101"
        assert ws.cell(2, monthly_col).value == 1000.0

    def test_export_multiple_rows(self, tmp_path):
        rows = [
            TenancyRow(property="Building A", tenant_name="ABC Corp"),
            TenancyRow(property="Building B", tenant_name="XYZ Inc"),
            TenancyRow(property="Building C", tenant_name="Test LLC"),
        ]
        output_path = tmp_path / "test.xlsx"
        export_tenancy_to_excel(rows, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Should have 1 header + 3 data rows
        assert ws.max_row == 4

    def test_export_without_warnings_column(self, tmp_path):
        rows = [
            TenancyRow(property="Test", warnings=["Some warning"]),
        ]
        output_path = tmp_path / "test.xlsx"
        export_tenancy_to_excel(rows, output_path, include_warnings=False)

        wb = load_workbook(output_path)
        ws = wb.active

        # Check that warnings column is not present
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        assert "warnings" not in headers

    def test_export_freeze_panes(self, tmp_path):
        rows = [
            TenancyRow(property="Test"),
        ]
        output_path = tmp_path / "test.xlsx"
        export_tenancy_to_excel(rows, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Panes should be frozen below header row
        assert ws.freeze_panes is not None

    def test_export_column_widths_set(self, tmp_path):
        rows = [
            TenancyRow(property="Test"),
        ]
        output_path = tmp_path / "test.xlsx"
        export_tenancy_to_excel(rows, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Check that column widths are set (not default)
        # Default width is typically around 8-10
        col_a_width = ws.column_dimensions["A"].width
        assert col_a_width > 10  # Should be wider than default


class TestExportTenancyToHtml:
    """Test HTML table export with reasoning block."""

    def _make_rows(self):
        return [
            TenancyRow(
                property="Cornet Axol",
                as_of_date="2024-09-30",
                tenant_name="Horizon Builders, LLC",
                suite="101",
                lease_from="2022-04-01",
                lease_to="2027-03-30",
                area_sqft=94940.0,
                monthly_amount=7911.67,
                annual_amount=94940.0,
                row_type="lease_summary",
            ),
            TenancyRow(
                property="Cornet Axol",
                as_of_date="2024-09-30",
                tenant_name="Horizon Builders, LLC",
                charge_label="CAM",
                period_from="2024-01-01",
                period_to="2024-12-31",
                monthly_amount=500.0,
                annual_amount=6000.0,
                management_fee_rate=0.05,
                row_type="charge_schedule",
            ),
        ]

    def test_returns_dict_with_required_keys(self):
        result = export_tenancy_to_html(self._make_rows())
        assert isinstance(result, dict)
        assert "html_table" in result
        assert "reasoning" in result

    def test_reasoning_has_required_sub_keys(self):
        result = export_tenancy_to_html(self._make_rows())
        reasoning = result["reasoning"]
        assert "parsing_strategy" in reasoning
        assert "normalization_decisions" in reasoning
        assert "warnings" in reasoning

    def test_html_table_is_string(self):
        result = export_tenancy_to_html(self._make_rows())
        assert isinstance(result["html_table"], str)

    def test_html_table_has_table_tags(self):
        result = export_tenancy_to_html(self._make_rows())
        html_table = result["html_table"]
        assert html_table.strip().startswith("<table>")
        assert html_table.strip().endswith("</table>")

    def test_html_table_has_thead_and_tbody(self):
        result = export_tenancy_to_html(self._make_rows())
        html_table = result["html_table"]
        assert "<thead>" in html_table
        assert "</thead>" in html_table
        assert "<tbody>" in html_table
        assert "</tbody>" in html_table

    def test_html_table_has_all_schema_columns_in_header(self):
        result = export_tenancy_to_html(self._make_rows())
        html_table = result["html_table"]
        for col in HTML_SCHEMA_COLUMNS:
            assert f"<th>{col}</th>" in html_table

    def test_html_schema_columns_order(self):
        """HTML_SCHEMA_COLUMNS must match the spec-mandated order."""
        required = [
            "property", "as_of_date", "row_type", "tenant_name", "suite",
            "lease_from", "lease_to", "area_sqft", "charge_label",
            "period_from", "period_to", "monthly_amount", "annual_amount",
            "management_fee_rate", "notes",
        ]
        assert HTML_SCHEMA_COLUMNS == required

    def test_html_table_has_data_rows(self):
        result = export_tenancy_to_html(self._make_rows())
        html_table = result["html_table"]
        assert html_table.count("<tr>") >= 3  # header + 2 data rows

    def test_html_table_data_in_correct_columns(self):
        rows = [TenancyRow(
            property="Test Property",
            tenant_name="Test Tenant",
            suite="202",
            row_type="lease_summary",
        )]
        result = export_tenancy_to_html(rows)
        html_table = result["html_table"]
        assert "Test Property" in html_table
        assert "Test Tenant" in html_table
        assert "202" in html_table
        assert "lease_summary" in html_table

    def test_html_table_has_td_cells_per_row(self):
        result = export_tenancy_to_html(self._make_rows())
        html_table = result["html_table"]
        # Each data row must contain exactly len(HTML_SCHEMA_COLUMNS) <td> cells
        import re
        data_rows = re.findall(r"<tr>(.*?)</tr>", html_table, re.DOTALL)
        # First <tr> is the header row (has <th> not <td>)
        for data_row in data_rows[1:]:
            td_count = data_row.count("<td>")
            assert td_count == len(HTML_SCHEMA_COLUMNS)

    def test_charge_label_and_period_fields_exported(self):
        result = export_tenancy_to_html(self._make_rows())
        html_table = result["html_table"]
        assert "CAM" in html_table
        assert "2024-01-01" in html_table
        assert "2024-12-31" in html_table

    def test_management_fee_rate_exported(self):
        result = export_tenancy_to_html(self._make_rows())
        html_table = result["html_table"]
        assert "0.05" in html_table

    def test_normalization_decisions_is_list(self):
        result = export_tenancy_to_html(self._make_rows())
        assert isinstance(result["reasoning"]["normalization_decisions"], list)
        assert len(result["reasoning"]["normalization_decisions"]) > 0

    def test_warnings_is_list(self):
        result = export_tenancy_to_html(self._make_rows())
        assert isinstance(result["reasoning"]["warnings"], list)

    def test_row_warnings_propagated(self):
        rows = [TenancyRow(
            property="Test",
            warnings=["Date '4/90/2026' interpreted as '2026-04-30'"],
        )]
        result = export_tenancy_to_html(rows)
        assert "4/90/2026" in result["reasoning"]["warnings"][0]

    def test_extra_warnings_appended(self):
        rows = [TenancyRow(property="Test")]
        result = export_tenancy_to_html(rows, warnings_list=["Extra warning"])
        assert "Extra warning" in result["reasoning"]["warnings"]

    def test_html_escaping(self):
        rows = [TenancyRow(
            property="<Acme & Sons>",
            tenant_name='Say "hello"',
        )]
        result = export_tenancy_to_html(rows)
        html_table = result["html_table"]
        assert "&lt;Acme &amp; Sons&gt;" in html_table
        assert "&quot;hello&quot;" in html_table

    def test_writes_json_file_when_output_path_given(self, tmp_path):
        output_path = tmp_path / "output.json"
        export_tenancy_to_html(self._make_rows(), output_path=output_path)
        assert output_path.exists()
        data = json.loads(output_path.read_text(encoding="utf-8"))
        assert "html_table" in data
        assert "reasoning" in data

    def test_json_file_is_valid_json(self, tmp_path):
        output_path = tmp_path / "output.json"
        export_tenancy_to_html(self._make_rows(), output_path=output_path)
        content = output_path.read_text(encoding="utf-8")
        parsed = json.loads(content)
        assert isinstance(parsed["html_table"], str)

    def test_empty_rows_produces_valid_html(self):
        result = export_tenancy_to_html([])
        html_table = result["html_table"]
        assert "<table>" in html_table
        assert "<thead>" in html_table
        assert "<tbody>" in html_table
        for col in HTML_SCHEMA_COLUMNS:
            assert f"<th>{col}</th>" in html_table

    def test_row_type_column_present_in_each_row(self):
        rows = [
            TenancyRow(property="P", row_type="rent_step"),
            TenancyRow(property="P", row_type="charge_schedule"),
        ]
        result = export_tenancy_to_html(rows)
        html_table = result["html_table"]
        assert "rent_step" in html_table
        assert "charge_schedule" in html_table


class TestDetectSectionType:
    """Test section-header detection for row_type assignment."""

    def test_rent_steps_header(self):
        assert _detect_section_type("Rent Steps") == ROW_TYPE_RENT_STEP

    def test_rent_steps_case_insensitive(self):
        assert _detect_section_type("RENT STEPS") == ROW_TYPE_RENT_STEP
        assert _detect_section_type("rent steps") == ROW_TYPE_RENT_STEP

    def test_rent_step_singular(self):
        assert _detect_section_type("Rent Step") == ROW_TYPE_RENT_STEP

    def test_rent_steps_with_noise(self):
        assert _detect_section_type("  Rent  Steps  ") == ROW_TYPE_RENT_STEP

    def test_charge_schedule_header(self):
        assert _detect_section_type("Charge Schedule") == ROW_TYPE_CHARGE_SCHEDULE

    def test_charge_schedules_plural(self):
        assert _detect_section_type("Charge Schedules") == ROW_TYPE_CHARGE_SCHEDULE

    def test_charge_schedule_case_insensitive(self):
        assert _detect_section_type("CHARGE SCHEDULE") == ROW_TYPE_CHARGE_SCHEDULE

    def test_occupancy_summary_header(self):
        assert _detect_section_type("Occupancy Summary") == ROW_TYPE_OCCUPANCY_SUMMARY

    def test_occupancy_summary_case_insensitive(self):
        assert _detect_section_type("OCCUPANCY SUMMARY") == ROW_TYPE_OCCUPANCY_SUMMARY

    def test_regular_data_row_returns_none(self):
        assert _detect_section_type("Horizon Builders, LLC") is None
        assert _detect_section_type("4/1/2022 3/30/2027 94940.00") is None

    def test_empty_string_returns_none(self):
        assert _detect_section_type("") is None

    def test_none_returns_none(self):
        assert _detect_section_type(None) is None


class TestExtractPropertyAsOfDate:
    """Test property name and as-of date extraction from document headers."""

    def test_extract_property_and_date(self):
        prop, aod = _extract_property_as_of_date(
            "Tenancy Schedule I Property: Cornet Axol Date: 09/30/2024"
        )
        assert prop == "Cornet Axol"
        assert aod == "2024-09-30"

    def test_extract_property_only(self):
        prop, aod = _extract_property_as_of_date("Property: Some Building Name")
        assert prop == "Some Building Name"
        assert aod is None

    def test_extract_date_only(self):
        prop, aod = _extract_property_as_of_date("As of: 2024-09-30")
        assert prop is None
        assert aod == "2024-09-30"

    def test_as_of_keyword(self):
        prop, aod = _extract_property_as_of_date(
            "Property: Test Corp As of: 01/15/2024"
        )
        assert prop == "Test Corp"
        assert aod == "2024-01-15"

    def test_empty_string(self):
        prop, aod = _extract_property_as_of_date("")
        assert prop is None
        assert aod is None

    def test_no_property_or_date(self):
        prop, aod = _extract_property_as_of_date("Horizon Builders 94940.00")
        assert prop is None
        assert aod is None


class TestParseGridRowTypeDetection:
    """Test that parse_grid_to_rows correctly assigns row_type from section headers."""

    def _make_section_grid(self) -> TableGrid:
        """Create a grid that simulates a tenancy schedule with multiple sections."""
        cells = [
            # Document header (row 0)
            CellRegion(row=0, col=0, text="Property: Test Plaza Date: 01/15/2024"),
            # Column headers (row 1)
            CellRegion(row=1, col=0, text="Tenant"),
            CellRegion(row=1, col=1, text="Suite"),
            CellRegion(row=1, col=2, text="Monthly"),
            CellRegion(row=1, col=3, text="Annual"),
            # Lease summary data (row 2)
            CellRegion(row=2, col=0, text="ABC Corp"),
            CellRegion(row=2, col=1, text="101"),
            CellRegion(row=2, col=2, text="1000.00"),
            CellRegion(row=2, col=3, text="12000.00"),
            # Rent Steps section header (row 3)
            CellRegion(row=3, col=0, text="Rent Steps"),
            # Rent step data (row 4)
            CellRegion(row=4, col=0, text="ABC Corp"),
            CellRegion(row=4, col=1, text="101"),
            CellRegion(row=4, col=2, text="1000.00"),
            CellRegion(row=4, col=3, text="12000.00"),
            # Charge Schedule section header (row 5)
            CellRegion(row=5, col=0, text="Charge Schedule"),
            # Charge schedule data (row 6)
            CellRegion(row=6, col=0, text="ABC Corp"),
            CellRegion(row=6, col=1, text="101"),
            CellRegion(row=6, col=2, text="50.00"),
            CellRegion(row=6, col=3, text="600.00"),
            # Occupancy Summary section header (row 7)
            CellRegion(row=7, col=0, text="Occupancy Summary"),
            # Occupancy data (row 8)
            CellRegion(row=8, col=0, text="OCCUPIED_TOTAL"),
            CellRegion(row=8, col=1, text="101"),
            CellRegion(row=8, col=2, text="500.00"),
            CellRegion(row=8, col=3, text="6000.00"),
        ]
        return TableGrid(cells=cells, header_rows=2)

    def test_lease_summary_rows(self):
        grid = self._make_section_grid()
        rows = parse_grid_to_rows(grid)
        lease_rows = [r for r in rows if r.row_type == ROW_TYPE_LEASE_SUMMARY]
        assert len(lease_rows) >= 1

    def test_rent_step_rows(self):
        grid = self._make_section_grid()
        rows = parse_grid_to_rows(grid)
        rent_rows = [r for r in rows if r.row_type == ROW_TYPE_RENT_STEP]
        assert len(rent_rows) >= 1

    def test_charge_schedule_rows(self):
        grid = self._make_section_grid()
        rows = parse_grid_to_rows(grid)
        charge_rows = [r for r in rows if r.row_type == ROW_TYPE_CHARGE_SCHEDULE]
        assert len(charge_rows) >= 1

    def test_occupancy_summary_rows(self):
        grid = self._make_section_grid()
        rows = parse_grid_to_rows(grid)
        occ_rows = [r for r in rows if r.row_type == ROW_TYPE_OCCUPANCY_SUMMARY]
        assert len(occ_rows) >= 1

    def test_all_four_row_types_present(self):
        grid = self._make_section_grid()
        rows = parse_grid_to_rows(grid)
        types = {r.row_type for r in rows}
        assert ROW_TYPE_LEASE_SUMMARY in types
        assert ROW_TYPE_RENT_STEP in types
        assert ROW_TYPE_CHARGE_SCHEDULE in types
        assert ROW_TYPE_OCCUPANCY_SUMMARY in types

    def test_property_propagated_from_header(self):
        grid = self._make_section_grid()
        rows = parse_grid_to_rows(grid)
        # All rows should have property set (propagated from doc header)
        assert all(r.property == "Test Plaza" for r in rows)

    def test_as_of_date_propagated_from_header(self):
        grid = self._make_section_grid()
        rows = parse_grid_to_rows(grid)
        # All rows should have as_of_date set (propagated from doc header)
        assert all(r.as_of_date == "2024-01-15" for r in rows)

    def test_section_header_rows_not_included_in_output(self):
        """Section header rows (Rent Steps etc.) must not appear as data rows."""
        grid = self._make_section_grid()
        rows = parse_grid_to_rows(grid)
        for row in rows:
            # tenant_name should never be the section header text
            assert row.tenant_name not in ("Rent Steps", "Charge Schedule", "Occupancy Summary")
