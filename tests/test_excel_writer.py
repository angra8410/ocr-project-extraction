"""Unit tests for the Excel writer module."""

import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from ocr_extractor.excel_writer import ExcelWriteError, _flag_low_confidence, write_xlsx
from ocr_extractor.table_detector import CellRegion, TableGrid


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _simple_grid(rows: int = 3, cols: int = 3, header_rows: int = 1) -> TableGrid:
    """Build a simple grid with sequential text content."""
    cells = []
    for r in range(rows):
        for c in range(cols):
            cells.append(
                CellRegion(
                    row=r,
                    col=c,
                    text=f"R{r}C{c}",
                    bbox=(c * 80, r * 40, (c + 1) * 80, (r + 1) * 40),
                )
            )
    return TableGrid(cells=cells, header_rows=header_rows)


# ---------------------------------------------------------------------------
# write_xlsx
# ---------------------------------------------------------------------------


class TestWriteXlsx:
    def test_creates_file(self, tmp_path):
        grid = _simple_grid()
        out = write_xlsx(grid, tmp_path / "out.xlsx")
        assert out.exists()

    def test_sheet_named_table(self, tmp_path):
        grid = _simple_grid()
        out = write_xlsx(grid, tmp_path / "out.xlsx")
        wb = load_workbook(str(out))
        assert "Table" in wb.sheetnames

    def test_one_sheet_only(self, tmp_path):
        grid = _simple_grid()
        out = write_xlsx(grid, tmp_path / "out.xlsx")
        wb = load_workbook(str(out))
        assert len(wb.sheetnames) == 1

    def test_cell_values_correct(self, tmp_path):
        grid = _simple_grid(rows=2, cols=2)
        out = write_xlsx(grid, tmp_path / "out.xlsx")
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.cell(1, 1).value == "R0C0"
        assert ws.cell(1, 2).value == "R0C1"
        assert ws.cell(2, 1).value == "R1C0"

    def test_empty_grid_creates_file(self, tmp_path):
        grid = TableGrid()
        out = write_xlsx(grid, tmp_path / "empty.xlsx")
        assert out.exists()

    def test_freeze_panes_below_header(self, tmp_path):
        grid = _simple_grid(rows=4, cols=3, header_rows=2)
        out = write_xlsx(grid, tmp_path / "out.xlsx")
        wb = load_workbook(str(out))
        ws = wb.active
        # Freeze pane should be at row 3 (below 2 header rows)
        assert ws.freeze_panes == "A3"

    def test_merged_header_cell(self, tmp_path):
        """A colspan=2 header cell should produce a merged region in xlsx."""
        cells = [
            CellRegion(row=0, col=0, colspan=2, text="Header", bbox=(0, 0, 160, 40)),
            CellRegion(row=0, col=2, text="H3", bbox=(160, 0, 240, 40)),
            CellRegion(row=1, col=0, text="A", bbox=(0, 40, 80, 80)),
            CellRegion(row=1, col=1, text="B", bbox=(80, 40, 160, 80)),
            CellRegion(row=1, col=2, text="C", bbox=(160, 40, 240, 80)),
        ]
        grid = TableGrid(cells=cells, header_rows=1)
        out = write_xlsx(grid, tmp_path / "merged.xlsx")
        wb = load_workbook(str(out))
        ws = wb.active
        # Check that A1:B1 is merged
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        assert any("A1" in mr and "B1" in mr for mr in merged_ranges)

    def test_numeric_values_right_aligned(self, tmp_path):
        cells = [
            CellRegion(row=0, col=0, text="Amount", bbox=(0, 0, 80, 40)),
            CellRegion(row=1, col=0, text="1,234.56", bbox=(0, 40, 80, 80)),
        ]
        grid = TableGrid(cells=cells, header_rows=1)
        out = write_xlsx(grid, tmp_path / "numeric.xlsx")
        wb = load_workbook(str(out))
        ws = wb.active
        data_cell = ws.cell(2, 1)
        assert data_cell.alignment.horizontal == "right"

    def test_low_confidence_flag_appended(self, tmp_path):
        cells = [
            CellRegion(row=0, col=0, text="Name"),
            CellRegion(row=1, col=0, text="John Doe", low_confidence=True),
        ]
        grid = TableGrid(cells=cells, header_rows=1)
        out = write_xlsx(grid, tmp_path / "lowconf.xlsx")
        wb = load_workbook(str(out))
        ws = wb.active
        value = ws.cell(2, 1).value
        assert value is not None
        assert "[?]" in str(value)

    def test_borders_present(self, tmp_path):
        grid = _simple_grid(rows=2, cols=2)
        out = write_xlsx(grid, tmp_path / "borders.xlsx")
        wb = load_workbook(str(out))
        ws = wb.active
        cell = ws.cell(1, 1)
        assert cell.border.left.border_style == "thin"

    def test_column_widths_set(self, tmp_path):
        grid = _simple_grid(rows=2, cols=3)
        out = write_xlsx(grid, tmp_path / "widths.xlsx")
        wb = load_workbook(str(out))
        ws = wb.active
        for col_letter in ("A", "B", "C"):
            assert ws.column_dimensions[col_letter].width > 0


# ---------------------------------------------------------------------------
# Atomic save / PermissionError
# ---------------------------------------------------------------------------


class TestAtomicSave:
    def test_no_tmp_file_left_on_success(self, tmp_path):
        """After a successful write the .tmp file must not remain."""
        grid = _simple_grid()
        out = tmp_path / "out.xlsx"
        write_xlsx(grid, out)
        tmp_file = tmp_path / "out.xlsx.tmp"
        assert not tmp_file.exists(), ".tmp file was not cleaned up after successful save"

    def test_permission_error_raises_excel_write_error(self, tmp_path):
        """A PermissionError from os.replace must become ExcelWriteError."""
        grid = _simple_grid()
        out = tmp_path / "locked.xlsx"
        with patch("os.replace", side_effect=PermissionError("Access denied")):
            with pytest.raises(ExcelWriteError, match="permission denied"):
                write_xlsx(grid, out)

    def test_excel_write_error_message_is_actionable(self, tmp_path):
        """ExcelWriteError message must mention closing Excel or choosing a path."""
        grid = _simple_grid()
        out = tmp_path / "locked.xlsx"
        with patch("os.replace", side_effect=PermissionError("Access denied")):
            with pytest.raises(ExcelWriteError) as exc_info:
                write_xlsx(grid, out)
        msg = str(exc_info.value)
        assert "Excel" in msg or "output path" in msg.lower() or "writable" in msg.lower()
