"""Integration / golden tests for the end-to-end extraction pipeline.

These tests run the full pipeline on a synthetic table image and verify
that the output .xlsx has the expected structure and content.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image, ImageDraw

from ocr_extractor.extractor import extract

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Fixture: synthetic table image
# ---------------------------------------------------------------------------

FIXTURES_DIR = Path(__file__).parent / "fixtures"
SAMPLE_TABLE = FIXTURES_DIR / "sample_table.png"
TEST_PDF = FIXTURES_DIR / "test.pdf"


def _make_table_image(
    content: list[list[str]],
    cell_w: int = 120,
    cell_h: int = 50,
    line_w: int = 2,
) -> Image.Image:
    """Draw a ruled-line table with the given cell content and return it."""
    rows = len(content)
    cols = max(len(r) for r in content)
    total_w = cols * cell_w + (cols + 1) * line_w
    total_h = rows * cell_h + (rows + 1) * line_w

    img = Image.new("RGB", (total_w, total_h), "white")
    draw = ImageDraw.Draw(img)

    for r in range(rows + 1):
        y = r * (cell_h + line_w)
        draw.rectangle([(0, y), (total_w, y + line_w)], fill="black")
    for c in range(cols + 1):
        x = c * (cell_w + line_w)
        draw.rectangle([(x, 0), (x + line_w, total_h)], fill="black")

    for r, row_data in enumerate(content):
        for c, text in enumerate(row_data):
            x = c * (cell_w + line_w) + line_w + 5
            y = r * (cell_h + line_w) + line_w + 15
            draw.text((x, y), text, fill="black")

    return img


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestExtractFromImage:
    def test_output_file_created(self, tmp_path):
        out = tmp_path / "out.xlsx"
        result = extract(str(SAMPLE_TABLE), str(out))
        assert Path(result).exists()

    def test_output_is_valid_xlsx(self, tmp_path):
        out = tmp_path / "out.xlsx"
        result = extract(str(SAMPLE_TABLE), str(out))
        wb = load_workbook(str(result))
        assert wb is not None

    def test_sheet_named_table(self, tmp_path):
        out = tmp_path / "out.xlsx"
        result = extract(str(SAMPLE_TABLE), str(out))
        wb = load_workbook(str(result))
        assert "Table" in wb.sheetnames

    def test_single_sheet_only(self, tmp_path):
        out = tmp_path / "out.xlsx"
        result = extract(str(SAMPLE_TABLE), str(out))
        wb = load_workbook(str(result))
        assert len(wb.sheetnames) == 1

    def test_has_multiple_rows(self, tmp_path):
        """The output should have at least the header + one data row."""
        out = tmp_path / "out.xlsx"
        result = extract(str(SAMPLE_TABLE), str(out))
        wb = load_workbook(str(result))
        ws = wb.active
        assert ws.max_row >= 2

    def test_has_multiple_columns(self, tmp_path):
        """The output should have at least 2 columns."""
        out = tmp_path / "out.xlsx"
        result = extract(str(SAMPLE_TABLE), str(out))
        wb = load_workbook(str(result))
        ws = wb.active
        assert ws.max_column >= 2

    def test_freeze_panes_set(self, tmp_path):
        """Panes should be frozen below the header row."""
        out = tmp_path / "out.xlsx"
        result = extract(str(SAMPLE_TABLE), str(out))
        wb = load_workbook(str(result))
        ws = wb.active
        assert ws.freeze_panes is not None

    def test_debug_mode_does_not_raise(self, tmp_path):
        """debug=True should complete without errors."""
        out = tmp_path / "debug_out.xlsx"
        result = extract(str(SAMPLE_TABLE), str(out), debug=True)
        assert Path(result).exists()


class TestExtractErrors:
    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            extract(str(tmp_path / "nonexistent.png"))

    def test_unsupported_extension_raises(self, tmp_path):
        f = tmp_path / "file.docx"
        f.write_bytes(b"dummy")
        with pytest.raises(ValueError, match="Unsupported"):
            extract(str(f))

    def test_default_output_path(self, tmp_path):
        """When output_path is omitted the .xlsx lands next to the input."""
        src = tmp_path / "table.png"
        import shutil

        shutil.copy(str(SAMPLE_TABLE), str(src))
        result = extract(str(src))
        assert Path(result).suffix == ".xlsx"
        assert Path(result).parent == tmp_path


class TestExtractSyntheticTable:
    """Golden test: verify OCR picks up expected column-header keywords."""

    def test_header_text_present(self, tmp_path):
        content = [
            ["Name", "Amount", "Date"],
            ["Alice", "100", "2024-01-01"],
            ["Bob", "200", "2024-01-02"],
        ]
        img = _make_table_image(content)
        src = tmp_path / "synthetic.png"
        img.save(str(src))
        out = tmp_path / "synthetic.xlsx"
        result = extract(str(src), str(out))
        wb = load_workbook(str(result))
        ws = wb.active

        # Collect all non-None cell values from the sheet
        all_values = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    all_values.append(str(v))

        all_text = " ".join(all_values).lower()

        # The header row must contain at least ONE of the known column names
        # (OCR may not be perfect, but at least something should come through)
        assert any(keyword in all_text for keyword in ("name", "amount", "date", "alice", "bob")), (
            f"No expected keywords found in extracted text. Got: {all_values}"
        )



# ---------------------------------------------------------------------------
# PDF integration tests (uses tests/fixtures/test.pdf)
# ---------------------------------------------------------------------------

# Skip the whole class if pdfplumber cannot open the fixture (e.g. in very
# restricted CI environments).  The PDF fixture is a PIL-generated single-page
# image-based PDF, so no special renderer is needed beyond pdfplumber.
_pdf_available = TEST_PDF.exists()


@pytest.mark.skipif(not _pdf_available, reason="tests/fixtures/test.pdf not found")
class TestPdfIntegration:
    """Integration tests that use tests/fixtures/test.pdf as input.

    Output is always written to pytest's tmp_path – never to the fixtures dir.
    """

    def test_pdf_output_written_to_tmp(self, tmp_path):
        """Output .xlsx must be created inside tmp_path, not in fixtures."""
        out = tmp_path / "test_output.xlsx"
        result = extract(str(TEST_PDF), str(out))
        result_path = Path(result)
        assert result_path.exists()
        # Crucially, the output must NOT be inside the fixtures directory
        assert not result_path.is_relative_to(FIXTURES_DIR), (
            f"Output was written to fixtures dir: {result_path}"
        )
        assert result_path.is_relative_to(tmp_path)

    def test_pdf_sheet_named_table(self, tmp_path):
        """The generated workbook must have exactly one sheet named 'Table'."""
        out = tmp_path / "test_output.xlsx"
        result = extract(str(TEST_PDF), str(out))
        wb = load_workbook(str(result))
        assert wb.sheetnames == ["Table"]

    def test_pdf_has_rows_and_columns(self, tmp_path):
        """The workbook must have at least 1 row and 1 column."""
        out = tmp_path / "test_output.xlsx"
        result = extract(str(TEST_PDF), str(out))
        wb = load_workbook(str(result))
        ws = wb.active
        assert ws.max_row >= 1
        assert ws.max_column >= 1

    def test_pdf_debug_artifacts_created(self, tmp_path):
        """Debug mode must create pipeline_diagram.md and grid_preview.txt."""
        out = tmp_path / "test_output.xlsx"
        result = extract(str(TEST_PDF), str(out), debug=True)
        debug_dir = tmp_path / "test_output.debug"
        assert debug_dir.exists(), f"Debug dir not found: {debug_dir}"
        diagram = debug_dir / "pipeline_diagram.md"
        preview = debug_dir / "grid_preview.txt"
        assert diagram.exists(), "pipeline_diagram.md not generated"
        assert preview.exists(), "grid_preview.txt not generated"
        # Check diagram has meaningful content
        content = diagram.read_text(encoding="utf-8")
        assert "Pipeline" in content
        assert "Header rows" in content or "header rows" in content.lower()
        # Check preview has meaningful content
        preview_content = preview.read_text(encoding="utf-8")
        assert "Grid Preview" in preview_content or "HDR" in preview_content

    def test_pdf_debug_artifacts_in_tmp_not_fixtures(self, tmp_path):
        """Debug artifacts must be written to tmp_path, not fixtures dir."""
        out = tmp_path / "test_output.xlsx"
        extract(str(TEST_PDF), str(out), debug=True)
        debug_dir = tmp_path / "test_output.debug"
        assert debug_dir.exists()
        assert not debug_dir.is_relative_to(FIXTURES_DIR)

    def test_pdf_multi_column_structure(self, tmp_path):
        """The output must have multi-column structure, NOT a single-column dump.
        
        This test fails if the extraction produces a single-column output where
        all data is dumped into column A (the "paragraph dump" problem).
        
        For tests/fixtures/test.pdf (real estate lease table), we expect:
        - At least 15 columns (17-column real estate lease data structure)
        - Header row with real estate column names (Property, Lease, Area, etc.)
        - Data rows with values in multiple columns
        """
        # Constants for this test - updated for 17-column real estate table
        MIN_EXPECTED_COLS = 15  # At least 15 of the 17 columns
        MIN_HEADER_VALUES = 8  # At least 8 distinct headers
        MIN_HEADER_COLS = 8    # Headers spread across at least 8 columns
        MIN_DATA_ROWS = 2      # At least 2 data rows with multi-column values
        MAX_COLS_TO_CHECK = 20  # Check up to 20 columns
        MAX_ROWS_TO_CHECK = 20  # Limit row checking for performance
        
        out = tmp_path / "test_output.xlsx"
        # Use tenancy_mode for structured multi-column output
        result = extract(str(TEST_PDF), str(out), tenancy_mode=True)
        wb = load_workbook(str(result))
        ws = wb.active
        
        # 1. Assert minimum column count for real estate table
        actual_cols = ws.max_column
        assert actual_cols >= MIN_EXPECTED_COLS, (
            f"Insufficient columns detected! Expected >= {MIN_EXPECTED_COLS} columns "
            f"for 17-column real estate table, but found only {actual_cols}. "
            f"This indicates the table structure was not properly detected."
        )
        
        # 2. Assert header values are in different columns (not all in A)
        header_row = 1
        header_values = []
        for col_idx in range(1, min(actual_cols + 1, MAX_COLS_TO_CHECK)):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value and str(cell_value).strip():
                header_values.append((col_idx, str(cell_value).strip()))
        
        # We expect real estate headers like "Property", "Lease", "Area", etc.
        expected_keywords = {"Property", "Lease", "Area", "Rent", "Annual", "Security"}
        found_keywords = set()
        for _, val in header_values:
            for keyword in expected_keywords:
                if keyword.lower() in val.lower():
                    found_keywords.add(keyword)
        
        assert len(header_values) >= MIN_HEADER_VALUES, (
            f"Too few header values found. Expected at least {MIN_HEADER_VALUES} distinct headers "
            f"for real estate table, found {len(header_values)}: {[h[1] for h in header_values[:10]]}"
        )
        
        assert len(found_keywords) >= 3, (
            f"Expected real estate header keywords not found. "
            f"Expected at least 3 of {expected_keywords}, found {found_keywords}. "
            f"All headers: {[h[1] for h in header_values[:10]]}"
        )
        
        # Check headers are in different columns (not all in column A)
        header_cols = {col for col, _ in header_values}
        assert len(header_cols) >= MIN_HEADER_COLS, (
            f"Headers are clustered in too few columns. Expected headers spread "
            f"across >= {MIN_HEADER_COLS} columns, but found in columns: {sorted(header_cols)}"
        )
        
        # 3. Assert at least N data rows have values beyond column A
        data_rows_with_multi_cols = 0
        
        for row_idx in range(2, min(ws.max_row + 1, MAX_ROWS_TO_CHECK)):
            has_value_beyond_A = False
            for col_idx in range(2, min(actual_cols + 1, MAX_COLS_TO_CHECK)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip():
                    has_value_beyond_A = True
                    break
            if has_value_beyond_A:
                data_rows_with_multi_cols += 1
        
        assert data_rows_with_multi_cols >= MIN_DATA_ROWS, (
            f"Too few data rows with multi-column values. Expected at least "
            f"{MIN_DATA_ROWS} rows with data in columns beyond A, but found "
            f"only {data_rows_with_multi_cols}. This suggests a single-column dump."
        )
        
        # 4. Check for expected merged cells (optional - test.pdf may not have merges)
        # For now, just document that we're not asserting merged cells
        merged_ranges = list(ws.merged_cells.ranges) if hasattr(ws, 'merged_cells') else []
        # Note: Not asserting merged cells for this fixture as it may not have them

    def test_pdf_generates_visual_artifacts(self, tmp_path):
        """Generate visual artifacts for debugging and verification.
        
        This test produces:
        - table_preview.html: Visual grid representation with styling
        - layout_overlay.png: Annotated image showing detected structure
        - test_assertions_report.md: Detailed explanation of what was checked
        
        These artifacts help verify that multi-column structure is preserved
        and provide visual evidence for code review.
        """
        from ocr_extractor.test_artifacts import (
            generate_assertions_report,
            generate_layout_overlay,
            generate_table_preview_html,
        )
        from ocr_extractor.preprocessor import preprocess
        from ocr_extractor.table_detector import detect_table, detect_merges
        from pdf2image import convert_from_path
        import cv2
        import numpy as np
        
        # Extract to Excel
        out = tmp_path / "test_output.xlsx"
        result = extract(str(TEST_PDF), str(out), debug=False)
        
        # Also get the image and grid for layout overlay
        images = convert_from_path(str(TEST_PDF), dpi=200)
        clean = preprocess(images[0], debug=False)
        
        # Detect table structure
        arr = cv2.cvtColor(np.array(clean.convert("RGB")), cv2.COLOR_RGB2BGR)
        gray = cv2.cvtColor(arr, cv2.COLOR_BGR2GRAY)
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        
        from ocr_extractor.table_detector import _detect_ruling_lines, _grid_from_lines
        h_lines, v_lines = _detect_ruling_lines(binary, debug=False)
        if h_lines is not None and v_lines is not None:
            grid = _grid_from_lines(h_lines, v_lines, gray.shape, debug=False)
        else:
            grid = detect_table(clean, debug=False)
        
        grid = detect_merges(clean, grid, debug=False)
        
        # Generate artifacts
        artifacts_dir = tmp_path / "test_artifacts"
        artifacts_dir.mkdir(exist_ok=True)
        
        # 1. Table preview HTML
        preview_html = artifacts_dir / "table_preview.html"
        generate_table_preview_html(
            xlsx_path=Path(result),
            output_path=preview_html,
            max_rows=30,
            max_cols=15,
        )
        assert preview_html.exists(), "table_preview.html not generated"
        
        # 2. Layout overlay PNG
        overlay_png = artifacts_dir / "layout_overlay.png"
        generate_layout_overlay(
            source_image=clean,
            grid=grid,
            output_path=overlay_png,
        )
        assert overlay_png.exists(), "layout_overlay.png not generated"
        
        # 3. Test assertions report
        report_md = artifacts_dir / "test_assertions_report.md"
        generate_assertions_report(
            xlsx_path=Path(result),
            output_path=report_md,
            test_name="PDF Multi-Column Structure Test",
        )
        assert report_md.exists(), "test_assertions_report.md not generated"
        
        # Verify content of report
        report_content = report_md.read_text(encoding="utf-8")
        assert "Column Structure Analysis" in report_content
        assert "Header Row Analysis" in report_content
        assert "Data Row Analysis" in report_content
        
        # Log locations for user
        logger.info("Visual artifacts generated:")
        logger.info("  - %s", preview_html)
        logger.info("  - %s", overlay_png)
        logger.info("  - %s", report_md)

    def test_pdf_golden_structure_expectations(self, tmp_path):
        """Verify the expected multi-column structure (golden expectations).
        
        This test encodes the INTENDED behavior for test.pdf (real estate lease table):
        - Expected headers: "Property", "Lease", "Area", "Rent", etc. in separate columns
        - Expected data: Property names, unit numbers, dates, financial values
        - This serves as a golden reference for the expected output
        """
        # Constants for this test
        MAX_ROWS_TO_CHECK = 20
        MAX_COLS_TO_CHECK = 20  # Check up to 20 columns for wide table
        MIN_PROPERTY_ENTRIES = 2  # At least 2 properties
        
        out = tmp_path / "test_output.xlsx"
        result = extract(str(TEST_PDF), str(out))
        wb = load_workbook(str(result))
        ws = wb.active
        
        # Golden expectations for real estate lease table
        
        # 1. Expected headers should exist in DIFFERENT columns
        expected_keywords = ["Property", "Lease", "Area", "Rent", "Annual"]
        
        # Find which columns contain these header keywords
        found_keywords = {}
        for keyword in expected_keywords:
            for col in range(1, min(ws.max_column + 1, MAX_COLS_TO_CHECK)):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and keyword.lower() in str(cell_value).lower():
                    found_keywords[keyword] = col
                    break
        
        # Assert at least 3 of the expected keywords were found
        assert len(found_keywords) >= 3, (
            f"Golden expectation failed: Expected at least 3 of {expected_keywords} "
            f"in headers, but only found {len(found_keywords)}: {found_keywords}"
        )
        
        # Assert keywords are in different columns (multi-column structure)
        unique_cols = set(found_keywords.values())
        assert len(unique_cols) >= 3, (
            f"Golden expectation failed: Header keywords should be in different columns, "
            f"but found in only {len(unique_cols)} columns: {sorted(unique_cols)}. "
            f"Mapping: {found_keywords}"
        )
        
        # 2. Expected data pattern: property names should appear
        expected_properties = ["AIP KKR", "PKP LKT", "KSN Southland", "Corner Fudge", "Precision"]
        found_properties = []
        
        # Search for property names in first few columns
        for row_idx in range(2, min(ws.max_row + 1, MAX_ROWS_TO_CHECK)):
            for col_idx in range(1, min(5, ws.max_column + 1)):  # Check first 5 columns
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    val_str = str(cell_value).strip()
                    for prop in expected_properties:
                        if prop.lower() in val_str.lower():
                            found_properties.append(prop)
                            break
        
        # At least 2 property entries should be found
        assert len(found_properties) >= MIN_PROPERTY_ENTRIES, (
            f"Golden expectation failed: Expected at least {MIN_PROPERTY_ENTRIES} property entries "
            f"from {expected_properties}, but only found {len(found_properties)}: {found_properties}"
        )
        
        # 3. Numeric values should appear in multiple columns (financial data)
        numeric_cols = 0
        for col_idx in range(1, min(ws.max_column + 1, MAX_COLS_TO_CHECK)):
            has_numeric = False
            for row_idx in range(2, min(ws.max_row + 1, 10)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    try:
                        val = float(str(cell_value).replace(",", ""))
                        if val > 0:  # Valid positive number
                            has_numeric = True
                            break
                    except (ValueError, TypeError):
                        pass
            if has_numeric:
                numeric_cols += 1
        
        assert numeric_cols >= 5, (
            f"Golden expectation failed: Expected at least 5 columns with numeric values "
            f"(financial data), but only found {numeric_cols}"
        )
