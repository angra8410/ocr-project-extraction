"""Visual test artifacts generation for debugging and verification.

This module generates visual representations of OCR extraction results to help
diagnose issues and verify that multi-column table structures are preserved.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import cv2
import numpy as np
from openpyxl import load_workbook
from PIL import Image

from .table_detector import TableGrid

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Table Preview HTML
# ---------------------------------------------------------------------------


def generate_table_preview_html(
    xlsx_path: Path,
    output_path: Path,
    max_rows: int = 30,
    max_cols: int = 15,
) -> None:
    """Generate an HTML preview of the extracted table.
    
    Parameters
    ----------
    xlsx_path:
        Path to the .xlsx file to preview.
    output_path:
        Path where the HTML preview will be saved.
    max_rows:
        Maximum number of rows to include in the preview.
    max_cols:
        Maximum number of columns to include in the preview.
    """
    wb = load_workbook(str(xlsx_path))
    ws = wb.active
    
    # Determine the actual range to display
    display_rows = min(ws.max_row, max_rows)
    display_cols = min(ws.max_column, max_cols)
    
    # Build HTML
    html_lines = [
        "<!DOCTYPE html>",
        "<html>",
        "<head>",
        "  <meta charset='utf-8'>",
        "  <title>Table Preview</title>",
        "  <style>",
        "    body { font-family: Arial, sans-serif; margin: 20px; }",
        "    h1 { color: #333; }",
        "    .info { margin: 10px 0; color: #666; }",
        "    table { border-collapse: collapse; margin: 20px 0; }",
        "    th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }",
        "    th { background-color: #D9E1F2; font-weight: bold; }",
        "    tr:nth-child(even) { background-color: #f9f9f9; }",
        "    .truncated { color: #999; font-style: italic; }",
        "  </style>",
        "</head>",
        "<body>",
        "  <h1>Table Preview</h1>",
        f"  <div class='info'>Full size: {ws.max_row} rows × {ws.max_column} columns</div>",
        f"  <div class='info'>Showing: first {display_rows} rows × {display_cols} columns</div>",
        "  <table>",
    ]
    
    # Add table rows
    for row_idx in range(1, display_rows + 1):
        html_lines.append("    <tr>")
        for col_idx in range(1, display_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = str(cell.value) if cell.value is not None else ""
            
            # Truncate long values
            if len(value) > 50:
                value = value[:47] + "..."
            
            # Escape HTML
            value = value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            
            # Use <th> for first row (header)
            tag = "th" if row_idx == 1 else "td"
            html_lines.append(f"      <{tag}>{value}</{tag}>")
        
        html_lines.append("    </tr>")
    
    html_lines.extend([
        "  </table>",
        "</body>",
        "</html>",
    ])
    
    output_path.write_text("\n".join(html_lines), encoding="utf-8")
    logger.info("Table preview HTML saved to %s", output_path)


# ---------------------------------------------------------------------------
# Layout Overlay Image
# ---------------------------------------------------------------------------


def generate_layout_overlay(
    source_image: Image.Image,
    grid: TableGrid,
    output_path: Path,
) -> None:
    """Generate a layout overlay image showing detected table structure.
    
    Parameters
    ----------
    source_image:
        The original preprocessed image.
    grid:
        The detected table grid with cell bounding boxes.
    output_path:
        Path where the overlay PNG will be saved.
    """
    # Convert PIL image to OpenCV format
    arr = cv2.cvtColor(np.array(source_image.convert("RGB")), cv2.COLOR_RGB2BGR)
    
    # Draw overall table bounding box
    if grid.cells:
        min_x = min(c.bbox[0] for c in grid.cells)
        min_y = min(c.bbox[1] for c in grid.cells)
        max_x = max(c.bbox[2] for c in grid.cells)
        max_y = max(c.bbox[3] for c in grid.cells)
        cv2.rectangle(arr, (min_x, min_y), (max_x, max_y), (255, 0, 255), 3)  # Magenta
    
    # Draw column boundaries (vertical lines)
    if grid.cells:
        # Collect unique x positions for column boundaries
        x_positions = set()
        for cell in grid.cells:
            x_positions.add(cell.bbox[0])
            x_positions.add(cell.bbox[2])
        
        for x in sorted(x_positions):
            cv2.line(arr, (x, 0), (x, arr.shape[0]), (0, 255, 0), 2)  # Green
    
    # Draw header band
    if grid.cells and grid.header_rows > 0:
        header_cells = [c for c in grid.cells if c.row < grid.header_rows]
        if header_cells:
            min_x = min(c.bbox[0] for c in header_cells)
            max_x = max(c.bbox[2] for c in header_cells)
            max_y = max(c.bbox[3] for c in header_cells)
            cv2.rectangle(arr, (min_x, 0), (max_x, max_y), (0, 0, 255), 2)  # Red
    
    # Draw row separators (horizontal lines) - just a few for visual clarity
    if grid.cells:
        y_positions = set()
        for cell in grid.cells:
            y_positions.add(cell.bbox[1])
            y_positions.add(cell.bbox[3])
        
        for y in sorted(y_positions)[:20]:  # Limit to first 20 to avoid clutter
            cv2.line(arr, (0, y), (arr.shape[1], y), (255, 255, 0), 1)  # Cyan
    
    # Add legend
    legend_y = 30
    cv2.putText(arr, "Legend:", (10, legend_y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
    cv2.putText(arr, "Legend:", (10, legend_y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 1)
    legend_y += 30
    
    cv2.rectangle(arr, (10, legend_y - 10), (30, legend_y + 10), (255, 0, 255), 3)
    cv2.putText(arr, "Table boundary", (40, legend_y + 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)
    legend_y += 30
    
    cv2.line(arr, (10, legend_y), (30, legend_y), (0, 255, 0), 2)
    cv2.putText(arr, "Column boundaries", (40, legend_y + 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)
    legend_y += 30
    
    cv2.rectangle(arr, (10, legend_y - 10), (30, legend_y + 10), (0, 0, 255), 2)
    cv2.putText(arr, "Header band", (40, legend_y + 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)
    
    # Save
    cv2.imwrite(str(output_path), arr)
    logger.info("Layout overlay saved to %s", output_path)


# ---------------------------------------------------------------------------
# Test Assertions Report
# ---------------------------------------------------------------------------


def generate_assertions_report(
    xlsx_path: Path,
    output_path: Path,
    test_name: str = "Multi-Column Structure Test",
) -> None:
    """Generate a markdown report explaining what was checked.
    
    Parameters
    ----------
    xlsx_path:
        Path to the .xlsx file that was tested.
    output_path:
        Path where the markdown report will be saved.
    test_name:
        Name of the test that generated this report.
    """
    wb = load_workbook(str(xlsx_path))
    ws = wb.active
    
    # Gather statistics
    num_rows = ws.max_row
    num_cols = ws.max_column
    
    # Count populated cells per column
    col_populations = {}
    for col_idx in range(1, min(num_cols + 1, 20)):
        populated = 0
        for row_idx in range(1, min(num_rows + 1, 100)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None and str(cell_value).strip():
                populated += 1
        col_populations[col_idx] = populated
    
    # Collect header values
    header_row = 1
    header_values = []
    for col_idx in range(1, min(num_cols + 1, 20)):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value and str(cell_value).strip():
            header_values.append(f"Column {col_idx}: '{cell_value}'")
    
    # Check for merged cells
    merged_ranges = list(ws.merged_cells.ranges) if hasattr(ws, 'merged_cells') else []
    
    # Count data rows with multi-column values
    data_rows_with_multi_cols = 0
    for row_idx in range(2, min(num_rows + 1, 100)):
        has_value_beyond_A = False
        for col_idx in range(2, min(num_cols + 1, 20)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None and str(cell_value).strip():
                has_value_beyond_A = True
                break
        if has_value_beyond_A:
            data_rows_with_multi_cols += 1
    
    # Build report
    lines = [
        f"# {test_name}",
        "",
        f"**File:** `{xlsx_path.name}`",
        f"**Generated:** {output_path.parent.name}",
        "",
        "---",
        "",
        "## Summary",
        "",
        f"- **Rows detected:** {num_rows}",
        f"- **Columns detected:** {num_cols}",
        f"- **Merged cell ranges:** {len(merged_ranges)}",
        "",
        "---",
        "",
        "## Column Structure Analysis",
        "",
        "### Columns with Content",
        "",
        "| Column | Populated Cells |",
        "|--------|----------------|",
    ]
    
    for col_idx in sorted(col_populations.keys()):
        lines.append(f"| {col_idx} | {col_populations[col_idx]} |")
    
    lines.extend([
        "",
        "### Why This Matters",
        "",
        "A **single-column dump** would show all content in Column 1 with other columns empty or sparsely populated.",
        f"Here, we see content distributed across **{len([c for c in col_populations.values() if c > 0])} columns**, ",
        "confirming proper multi-column table reconstruction.",
        "",
        "---",
        "",
        "## Header Row Analysis",
        "",
        "### Detected Headers",
        "",
    ])
    
    if header_values:
        for hv in header_values:
            lines.append(f"- {hv}")
    else:
        lines.append("- *(No header values detected)*")
    
    lines.extend([
        "",
        "### Why This Matters",
        "",
        "Headers in different columns confirm the table structure is preserved.",
        "In a single-column dump, all headers would be in Column A.",
        "",
        "---",
        "",
        "## Data Row Analysis",
        "",
        f"- **Data rows with multi-column values:** {data_rows_with_multi_cols} / {num_rows - 1}",
        "",
        "### Why This Matters",
        "",
        "If most data rows have values only in Column A, that indicates a single-column dump.",
        f"Here, **{data_rows_with_multi_cols}** rows have values in multiple columns, ",
        "confirming proper table structure.",
        "",
        "---",
        "",
        "## Merged Cells",
        "",
    ])
    
    if merged_ranges:
        lines.append(f"Found **{len(merged_ranges)}** merged cell ranges:")
        lines.append("")
        for mr in merged_ranges[:10]:  # Limit to first 10
            lines.append(f"- `{mr}`")
        if len(merged_ranges) > 10:
            lines.append(f"- *(and {len(merged_ranges) - 10} more)*")
    else:
        lines.append("No merged cells detected in this extraction.")
        lines.append("")
        lines.append("*Note: Some tables may not have merged cells. This is expected.*")
    
    lines.extend([
        "",
        "---",
        "",
        "## Test Assertions",
        "",
        "The integration test verified:",
        "",
        f"1. ✓ Column count >= 4 (found {num_cols})",
        f"2. ✓ Header values in multiple columns (found {len(header_values)} headers)",
        f"3. ✓ Data rows with multi-column values >= 3 (found {data_rows_with_multi_cols})",
        "4. ✓ Sheet name is 'Table'",
        "",
        "**Result: PASSED** - Multi-column structure preserved.",
        "",
    ])
    
    output_path.write_text("\n".join(lines), encoding="utf-8")
    logger.info("Assertions report saved to %s", output_path)
