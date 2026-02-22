"""Tenancy schedule parser for extracting structured data from real estate lease documents.

This module provides specialized parsing for tenancy schedules that contain
multiple tables and repeated sections like:
- Tenancy Schedule (main lease data)
- Rent Steps
- Charge Schedules
- Occupancy Summary

The parser handles:
- Column boundary detection for wide tables (15-20+ columns)
- Numeric normalization (remove commas, handle parentheses as negatives)
- Date normalization to ISO YYYY-MM-DD format
- Null handling for ambiguous data
- Warning tracking for uncertain extractions
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .table_detector import TableGrid

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Schema Definition
# ---------------------------------------------------------------------------

# Column definitions for tenancy schedule tables
TENANCY_SCHEDULE_COLUMNS = [
    "property",
    "as_of_date",
    "tenant_name",
    "legal_name",
    "suite",
    "lease_type",
    "lease_from",
    "lease_to",
    "term_months",
    "area_sqft",
    "monthly_amount",
    "annual_amount",
    "security_deposit",
    "loc_amount",
    "notes",
    "row_type",
    "warnings",
]

# Row type constants
ROW_TYPE_LEASE_SUMMARY = "lease_summary"
ROW_TYPE_RENT_STEP = "rent_step"
ROW_TYPE_CHARGE_SCHEDULE = "charge_schedule"
ROW_TYPE_OCCUPANCY_SUMMARY = "occupancy_summary"
ROW_TYPE_HEADER = "header"


# ---------------------------------------------------------------------------
# Data Models
# ---------------------------------------------------------------------------


@dataclass
class TenancyRow:
    """A single row in the tenancy schedule."""

    property: Optional[str] = None
    as_of_date: Optional[str] = None
    tenant_name: Optional[str] = None
    legal_name: Optional[str] = None
    suite: Optional[str] = None
    lease_type: Optional[str] = None
    lease_from: Optional[str] = None
    lease_to: Optional[str] = None
    term_months: Optional[float] = None
    area_sqft: Optional[float] = None
    monthly_amount: Optional[float] = None
    annual_amount: Optional[float] = None
    security_deposit: Optional[float] = None
    loc_amount: Optional[float] = None
    notes: Optional[str] = None
    row_type: str = ROW_TYPE_LEASE_SUMMARY
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary for DataFrame construction."""
        data = {
            "property": self.property,
            "as_of_date": self.as_of_date,
            "tenant_name": self.tenant_name,
            "legal_name": self.legal_name,
            "suite": self.suite,
            "lease_type": self.lease_type,
            "lease_from": self.lease_from,
            "lease_to": self.lease_to,
            "term_months": self.term_months,
            "area_sqft": self.area_sqft,
            "monthly_amount": self.monthly_amount,
            "annual_amount": self.annual_amount,
            "security_deposit": self.security_deposit,
            "loc_amount": self.loc_amount,
            "notes": self.notes,
            "row_type": self.row_type,
            "warnings": "; ".join(self.warnings) if self.warnings else None,
        }
        return data


# ---------------------------------------------------------------------------
# Normalization Functions
# ---------------------------------------------------------------------------


def normalize_number(value: str) -> Optional[float]:
    """Normalize numeric values from OCR text.

    Handles:
    - Removes commas (1,234.56 → 1234.56)
    - Handles parentheses as negatives: (100) → -100
    - Handles OCR errors: O vs 0
    - Returns None for invalid/ambiguous values

    Args:
        value: String value to normalize

    Returns:
        Normalized float or None if invalid
    """
    if not value or not isinstance(value, str):
        return None

    # Clean the value
    value = value.strip()

    if not value:
        return None

    # Handle parentheses as negative
    is_negative = False
    if value.startswith("(") and value.endswith(")"):
        is_negative = True
        value = value[1:-1].strip()

    # Remove commas and dollar signs
    value = value.replace(",", "").replace("$", "").strip()

    # Common OCR errors: replace O with 0
    # Only do this if the value looks like a number (contains digits)
    if re.search(r"\d", value):
        value = value.replace("O", "0").replace("o", "0")

    # Try to convert to float
    try:
        result = float(value)
        return -result if is_negative else result
    except ValueError:
        logger.debug("Cannot normalize number: %r", value)
        return None


def normalize_date(value: str) -> Optional[str]:
    """Normalize date values to ISO YYYY-MM-DD format.

    Handles common formats:
    - MM/DD/YYYY, M/D/YYYY
    - DD-MM-YYYY, D-M-YYYY
    - Month DD, YYYY
    - DD Month YYYY

    Args:
        value: String value to normalize

    Returns:
        ISO date string (YYYY-MM-DD) or None if invalid
    """
    if not value or not isinstance(value, str):
        return None

    value = value.strip()

    if not value:
        return None

    # Common date formats to try
    formats = [
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%B %d, %Y",
        "%b %d, %Y",
        "%d %B %Y",
        "%d %b %Y",
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(value, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    logger.debug("Cannot normalize date: %r", value)
    return None


# ---------------------------------------------------------------------------
# Grid-to-Row Parsing
# ---------------------------------------------------------------------------


def parse_grid_to_rows(grid: TableGrid) -> list[TenancyRow]:
    """Parse a TableGrid into structured tenancy rows.

    This function extracts data from the grid and converts it into
    structured TenancyRow objects with proper column mapping.

    Args:
        grid: TableGrid from table detection

    Returns:
        List of TenancyRow objects
    """
    if not grid.cells:
        logger.warning("Empty grid provided to parse_grid_to_rows")
        return []

    rows: list[TenancyRow] = []

    # Build a map: (row, col) -> cell for fast lookup
    cell_map: dict[tuple[int, int], Any] = {}
    for cell in grid.cells:
        cell_map[(cell.row, cell.col)] = cell

    # Detect header row and column mapping
    header_map = _detect_header_mapping(grid, cell_map)

    if not header_map:
        logger.warning("Could not detect header mapping; using fallback column order")
        # Fallback: assume columns in standard order
        header_map = _create_fallback_header_mapping(grid.num_cols)

    # Process data rows (skip header rows)
    for row_idx in range(grid.header_rows, grid.num_rows):
        tenancy_row = _extract_row_data(row_idx, grid.num_cols, cell_map, header_map)

        # Only add rows that have at least some data
        if _has_meaningful_data(tenancy_row):
            rows.append(tenancy_row)

    logger.info("Parsed %d tenancy rows from grid (%d rows × %d cols)",
                len(rows), grid.num_rows, grid.num_cols)

    return rows


def _detect_header_mapping(grid: TableGrid, cell_map: dict) -> dict[str, int]:
    """Detect which columns correspond to which fields based on header text.

    Args:
        grid: TableGrid
        cell_map: Map of (row, col) -> cell

    Returns:
        Dictionary mapping field names to column indices
    """
    header_map: dict[str, int] = {}

    # Keywords to look for in headers
    keywords = {
        "property": ["property", "building", "site"],
        "tenant_name": ["tenant", "name", "company"],
        "legal_name": ["legal", "legal name"],
        "suite": ["suite", "unit", "space"],
        "lease_type": ["type", "lease type"],
        "lease_from": ["from", "start", "commencement", "commence"],
        "lease_to": ["to", "end", "expiration", "expire", "expiry"],
        "term_months": ["term", "months", "duration"],
        "area_sqft": ["area", "sqft", "sf", "square"],
        "monthly_amount": ["monthly", "month", "rent/month"],
        "annual_amount": ["annual", "year", "rent/year"],
        "security_deposit": ["security", "deposit"],
        "loc_amount": ["loc", "letter", "credit"],
    }

    # Check first few rows for headers
    for row_idx in range(min(grid.header_rows, 3)):
        for col_idx in range(grid.num_cols):
            cell = cell_map.get((row_idx, col_idx))
            if not cell or not cell.text:
                continue

            cell_text = cell.text.lower().strip()

            # Match against keywords
            for field, kws in keywords.items():
                if field in header_map:
                    continue  # Already found

                for kw in kws:
                    if kw in cell_text:
                        header_map[field] = col_idx
                        logger.debug("Mapped %r to column %d (header: %r)",
                                     field, col_idx, cell.text)
                        break

    return header_map


def _create_fallback_header_mapping(num_cols: int) -> dict[str, int]:
    """Create a fallback header mapping when automatic detection fails.

    Assumes columns in this order:
    Property, Tenant, Suite, Lease From, Lease To, Area, Monthly, Annual, ...
    """
    fallback_order = [
        "property",
        "tenant_name",
        "suite",
        "lease_from",
        "lease_to",
        "area_sqft",
        "monthly_amount",
        "annual_amount",
        "security_deposit",
        "loc_amount",
        "notes",
    ]

    header_map = {}
    for idx, field in enumerate(fallback_order):
        if idx < num_cols:
            header_map[field] = idx

    return header_map


def _extract_row_data(
    row_idx: int,
    num_cols: int,
    cell_map: dict,
    header_map: dict[str, int],
) -> TenancyRow:
    """Extract data from a single row into a TenancyRow object.

    Args:
        row_idx: Row index to extract
        num_cols: Total number of columns
        cell_map: Map of (row, col) -> cell
        header_map: Mapping of field names to column indices

    Returns:
        TenancyRow object
    """
    tenancy_row = TenancyRow()

    # Helper to get cell text
    def get_cell_text(col_idx: int) -> str:
        cell = cell_map.get((row_idx, col_idx))
        return cell.text.strip() if cell and cell.text else ""

    # Extract text fields
    if "property" in header_map:
        tenancy_row.property = get_cell_text(header_map["property"]) or None

    if "tenant_name" in header_map:
        tenancy_row.tenant_name = get_cell_text(header_map["tenant_name"]) or None

    if "legal_name" in header_map:
        tenancy_row.legal_name = get_cell_text(header_map["legal_name"]) or None

    if "suite" in header_map:
        tenancy_row.suite = get_cell_text(header_map["suite"]) or None

    if "lease_type" in header_map:
        tenancy_row.lease_type = get_cell_text(header_map["lease_type"]) or None

    if "notes" in header_map:
        tenancy_row.notes = get_cell_text(header_map["notes"]) or None

    # Extract and normalize dates
    if "lease_from" in header_map:
        raw_from = get_cell_text(header_map["lease_from"])
        if raw_from:
            normalized = normalize_date(raw_from)
            if normalized:
                tenancy_row.lease_from = normalized
            else:
                tenancy_row.lease_from = raw_from
                tenancy_row.warnings.append(f"Could not parse date: {raw_from}")

    if "lease_to" in header_map:
        raw_to = get_cell_text(header_map["lease_to"])
        if raw_to:
            normalized = normalize_date(raw_to)
            if normalized:
                tenancy_row.lease_to = normalized
            else:
                tenancy_row.lease_to = raw_to
                tenancy_row.warnings.append(f"Could not parse date: {raw_to}")

    # Extract and normalize numeric fields
    numeric_fields = [
        ("term_months", "term_months"),
        ("area_sqft", "area_sqft"),
        ("monthly_amount", "monthly_amount"),
        ("annual_amount", "annual_amount"),
        ("security_deposit", "security_deposit"),
        ("loc_amount", "loc_amount"),
    ]

    for field, key in numeric_fields:
        if key in header_map:
            raw_value = get_cell_text(header_map[key])
            if raw_value:
                normalized = normalize_number(raw_value)
                if normalized is not None:
                    setattr(tenancy_row, field, normalized)
                else:
                    # Add warning for unparseable value
                    tenancy_row.warnings.append(f"Could not parse {field}: {raw_value}")

    return tenancy_row


def _has_meaningful_data(row: TenancyRow) -> bool:
    """Check if a row has any meaningful data (not all None/empty)."""
    return any([
        row.property,
        row.tenant_name,
        row.suite,
        row.lease_from,
        row.lease_to,
        row.area_sqft,
        row.monthly_amount,
        row.annual_amount,
    ])


# ---------------------------------------------------------------------------
# Excel Export with Multi-Column Guarantee
# ---------------------------------------------------------------------------


def export_tenancy_to_excel(
    rows: list[TenancyRow],
    output_path: str | Path,
    include_warnings: bool = True,
) -> Path:
    """Export tenancy rows to Excel with guaranteed multi-column structure.

    This function ensures that the output Excel file has:
    - Multiple columns with clear headers
    - Proper data types (numbers as numbers, dates as dates)
    - No single-column dumps

    Args:
        rows: List of TenancyRow objects
        output_path: Path to output .xlsx file
        include_warnings: Whether to include the warnings column

    Returns:
        Path to the created Excel file
    """
    output_path = Path(output_path)

    # Convert rows to dictionaries
    data = [row.to_dict() for row in rows]

    # Create DataFrame with explicit column order
    columns = TENANCY_SCHEDULE_COLUMNS.copy()
    if not include_warnings:
        columns.remove("warnings")

    df = pd.DataFrame(data, columns=columns)

    # Create workbook manually with openpyxl for better formatting control
    wb = Workbook()
    ws = wb.active
    ws.title = "Tenancy Schedule"

    # Write headers
    _THIN = Side(border_style="thin", color="000000")
    _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    _HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
    _HEADER_FONT = Font(bold=True)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    # Write data rows
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data.get(col_name)

            # Write value with appropriate type
            if value is not None:
                cell.value = value
            else:
                cell.value = ""

            cell.border = _BORDER

            # Right-align numeric columns
            if col_name in ["term_months", "area_sqft", "monthly_amount",
                            "annual_amount", "security_deposit", "loc_amount"]:
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Set column widths
    column_widths = {
        "property": 20,
        "as_of_date": 12,
        "tenant_name": 25,
        "legal_name": 25,
        "suite": 10,
        "lease_type": 12,
        "lease_from": 12,
        "lease_to": 12,
        "term_months": 10,
        "area_sqft": 12,
        "monthly_amount": 15,
        "annual_amount": 15,
        "security_deposit": 15,
        "loc_amount": 15,
        "notes": 30,
        "row_type": 15,
        "warnings": 40,
    }

    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = column_widths.get(col_name, 15)

    # Freeze header row
    ws.freeze_panes = ws.cell(row=2, column=1)

    # Save workbook
    wb.save(output_path)
    logger.info("Wrote tenancy schedule to %s (%d rows, %d columns)",
                output_path, len(rows), len(columns))

    return output_path.resolve()
