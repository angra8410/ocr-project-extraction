"""Excel (.xlsx) writer.

Takes a :class:`~ocr_extractor.table_detector.TableGrid` and writes a
single-sheet workbook that mimics the output style of jpgtoexcel.com:
- One sheet named ``Table``
- Merged cells for header spans
- Frozen panes below the header area
- Simple border on every cell
- Text left-aligned; numbers right-aligned
- IDs / dates kept as text to avoid auto-formatting
- Low-confidence cells flagged with a trailing ``[?]`` marker (or an
  Excel cell comment when openpyxl supports it)
"""

from __future__ import annotations

import logging
import os
import re
import tempfile
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .table_detector import CellRegion, TableGrid


class ExcelWriteError(OSError):
    """Raised when the workbook cannot be saved to *output_path*."""

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_THIN = Side(border_style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
_HEADER_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")

# Regex to decide whether a cell value looks "numeric"
_NUMERIC_RE = re.compile(r"^[+-]?[\d,]+(\.\d+)?$")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def write_xlsx(grid: TableGrid, output_path: str | Path) -> Path:
    """Write *grid* to a new .xlsx file at *output_path*.

    Parameters
    ----------
    grid:
        Populated :class:`TableGrid` (cells must have ``text`` filled in).
    output_path:
        Destination .xlsx path.  Parent directory must exist.

    Returns
    -------
    Path
        Resolved path of the written file.
    """
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"

    if not grid.cells:
        _atomic_save(wb, output_path)
        return output_path.resolve()

    num_cols = grid.num_cols

    # 1. Write cells (merged or plain)
    _write_cells(ws, grid)

    # 2. Freeze panes below the header
    freeze_row = grid.header_rows + 1  # openpyxl rows are 1-based
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)

    # 3. Set column widths proportional to bounding-box widths
    _set_column_widths(ws, grid, num_cols)

    _atomic_save(wb, output_path)
    logger.info("Wrote %s", output_path)
    return output_path.resolve()


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _atomic_save(wb: Workbook, output_path: Path) -> None:
    """Save *wb* to *output_path* using an atomic write-then-replace pattern.

    Writes to a sibling ``.tmp`` file first, then replaces the target.
    This prevents partial writes from corrupting an existing file.

    Raises :class:`ExcelWriteError` with an actionable message when the save
    fails due to a permission or locking problem.
    """
    tmp_path = output_path.parent / f"{output_path.name}.tmp"
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, output_path)
    except PermissionError as exc:
        # Clean up the temp file if it was created
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
        suggestions = [
            f"  • Close '{output_path.name}' in Excel if it is open.",
            f"  • Choose a different output path (current: {output_path}).",
            "  • Check that the directory is writable and not read-only.",
        ]
        raise ExcelWriteError(
            f"Cannot write '{output_path}': permission denied.\n"
            + "\n".join(suggestions)
        ) from exc
    except Exception:
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise


def _write_cells(ws, grid: TableGrid) -> None:  # type: ignore[type-arg]
    """Write all cells from *grid* into worksheet *ws*."""
    # Track which (xl_row, xl_col) positions are already occupied by a merge
    occupied: set[tuple[int, int]] = set()

    # Sort cells top-left to bottom-right for deterministic writing order
    sorted_cells = sorted(grid.cells, key=lambda c: (c.row, c.col))

    for cell in sorted_cells:
        xl_row = cell.row + 1  # 1-based
        xl_col = cell.col + 1  # 1-based

        if (xl_row, xl_col) in occupied:
            continue

        text = cell.text
        is_header = cell.row < grid.header_rows

        # Determine cell value (keep text as text; optionally right-align numbers)
        value: str | float | int
        is_numeric = bool(_NUMERIC_RE.match(text.replace(",", ""))) and not is_header
        if is_numeric and text:
            try:
                # Strip thousands separators, keep as float/int
                cleaned = text.replace(",", "")
                value = int(cleaned) if "." not in cleaned else float(cleaned)
            except ValueError:
                value = text
        else:
            value = text

        xl_cell = ws.cell(row=xl_row, column=xl_col, value=value)
        xl_cell.border = _BORDER

        # Alignment
        if is_numeric and not is_header:
            xl_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="right")
        else:
            xl_cell.alignment = _WRAP

        # Header styling
        if is_header:
            xl_cell.font = _HEADER_FONT
            xl_cell.fill = _HEADER_FILL

        # Low-confidence flag
        if cell.low_confidence:
            _flag_low_confidence(xl_cell, text)

        # Merged cells
        if cell.is_merged:
            end_row = xl_row + cell.rowspan - 1
            end_col = xl_col + cell.colspan - 1
            ws.merge_cells(
                start_row=xl_row,
                start_column=xl_col,
                end_row=end_row,
                end_column=end_col,
            )
            # Mark occupied positions so we don't write over them
            for r in range(xl_row, end_row + 1):
                for c in range(xl_col, end_col + 1):
                    if (r, c) != (xl_row, xl_col):
                        occupied.add((r, c))


def _flag_low_confidence(xl_cell, original_text: str) -> None:
    """Append ``[?]`` to the cell value and add an Excel comment."""
    current = xl_cell.value
    if isinstance(current, str) and not current.endswith("[?]"):
        xl_cell.value = current + " [?]"
    try:
        xl_cell.comment = Comment(
            text=(
                f"Low-confidence OCR. Original detected text: {original_text!r}\n"
                "Please verify this value."
            ),
            author="OCR Extractor",
        )
    except Exception:  # noqa: BLE001
        pass  # Comments are best-effort


def _set_column_widths(ws, grid: TableGrid, num_cols: int) -> None:  # type: ignore[type-arg]
    """Set column widths based on bbox widths or text length heuristic."""
    # Try to derive widths from cell bounding boxes
    col_pixel_widths: list[float] = [0.0] * num_cols
    for cell in grid.cells:
        if cell.bbox != (0, 0, 0, 0):
            pixel_width = cell.bbox[2] - cell.bbox[0]
            # Distribute evenly over colspan
            per_col = pixel_width / cell.colspan
            for c in range(cell.col, cell.col + cell.colspan):
                if c < num_cols:
                    col_pixel_widths[c] = max(col_pixel_widths[c], per_col)

    total_px = sum(col_pixel_widths)

    # Also consider text length as a lower bound
    col_text_widths: list[int] = [8] * num_cols
    for cell in grid.cells:
        if cell.text and cell.colspan == 1:
            # Estimate character width (max line length in cell text)
            max_line = max((len(line) for line in cell.text.split("\n")), default=0)
            col_text_widths[cell.col] = max(col_text_widths[cell.col], max_line + 2)

    for col_idx in range(num_cols):
        xl_col = col_idx + 1
        col_letter = get_column_letter(xl_col)

        if total_px > 0:
            # Proportional width scaled to a ~120-character total width
            proportional = (col_pixel_widths[col_idx] / total_px) * 120
            width = max(proportional, float(col_text_widths[col_idx]))
        else:
            width = float(col_text_widths[col_idx])

        ws.column_dimensions[col_letter].width = min(width, 60)
